from enum import Flag
import re
import time
from unittest import result

import openpyxl
import pyodbc
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill
from tqdm import tqdm

from config import (DatabaseHost, DatabasePassword, Mode, smartPassword,
                    smartUserName)
from seleniumDriver import CreateEdgeDriverService
from smartLogin import SmartLogin


class Database:
    # 初始化数据库连接
    DatabaseName = 'PeopleDaily+'
    DateTableName = "DateIndex"
    contentTableName = "PaperIndex"

    def __init__(self):
        self.cursor = self.connetMsSqlServer()
        # self.cursor.execute("USE MASTER")   #Debug
        # self.cursor.execute("DROP DATABASE PeopleDaily")  # 删除旧数据库 #Debug
        self.CreateDatabase()
        self.CreateTable()

    def CreateDatabase(self):
        self.cursor.execute(
            f"select * From master.dbo.sysdatabases where name= '{self.DatabaseName}'")
        if self.cursor.fetchone() == None:
            self.cursor.execute(
                f"Create Database [{self.DatabaseName}]")
        self.cursor.execute(f"use [{self.DatabaseName}]")

    def CreateTable(self):
        DateTableName = self.DateTableName
        contentTableName = self.contentTableName
        SQL = f"""IF NOT EXISTS (SELECT * FROM sys.objects WHERE object_id = OBJECT_ID(N'[dbo].[{DateTableName}]') AND type in (N'U'))
                    CREATE TABLE [dbo].[{DateTableName}] (
                    [Date]           DATE           NOT NULL,
                    [PaperNumbers] INT            NULL,
                    [DailyURL]       NVARCHAR (MAX) NULL,
                    [Exported-PaperNumbers] INT  NULL,
                    PRIMARY KEY CLUSTERED ([Date] ASC)
                )""".replace("\n", " ")
        self.cursor.execute(SQL)
        SQL = f"""IF NOT EXISTS (SELECT * FROM sys.objects WHERE object_id = OBJECT_ID(N'[dbo].[{contentTableName}]') AND type in (N'U'))
                    CREATE TABLE [dbo].[{contentTableName}] (
                    [PaperIndex] NVARCHAR (20) NOT NULL,
                    [Date]         DATE           NOT NULL,
                    [PaperTitle] NVARCHAR (450) NOT NULL,
                    [Author]       NVARCHAR (MAX) NULL,
                    [Keyword]      NVARCHAR (MAX) NULL,
                    [Page]         NVARCHAR (MAX) NULL,
                    [PaperURL]   NVARCHAR (MAX) NOT NULL,
                    PRIMARY KEY CLUSTERED ([Date] ASC, [PaperTitle] ASC),
                    FOREIGN KEY ([Date]) REFERENCES [dbo].[{DateTableName}] ([Date])
                );""".replace("\n", " ")
        self.cursor.execute(SQL)
        try:
            CreateViewSQL = f"""CREATE VIEW [Date_Paper] AS (                            
                                SELECT A.[Date]                                 
                                ,[PaperNumbers]                                 
                                ,[DailyURL]                                
                                ,B.[Exported-PaperNumbers]                  
                                FROM [dbo].[DateIndex] A LEFT JOIN     
                                (SELECT [Date], COUNT(*) AS [Exported-PaperNumbers] FROM [dbo].[PaperIndex] GROUP BY [Date]) B         
                                ON A.[Date] = B.[Date] )""".replace("\n", " ")
            self.cursor.execute(CreateViewSQL)
        except:
            pass

    def connetMsSqlServer(self):
        server = DatabaseHost
        username = 'sa'
        password = DatabasePassword
        DB = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
                            server + ';UID=' + username + ';PWD=' + password, autocommit=True)
        cursor = DB.cursor()
        return cursor

    def InsertData_Date(self, Date, PaperNumbers, DailyURL):
        if DailyURL != None:
            DailyURL = f"'{DailyURL}'"
        InsertSQL = f"INSERT INTO [{self.DateTableName}] VALUES ('{Date}',{PaperNumbers},{DailyURL},'')"
        try:
            self.cursor.execute(InsertSQL)
        except Exception as err:
            # 判断错误是否因主键冲突
            PRIMARYKEY_ERROR_CHECK = re.search('PRIMARY KEY', str(err))
            if PRIMARYKEY_ERROR_CHECK == None:
                raise
            else:
                PrintAndSave(f"主键错误：{Date}已存在")

    def InsertData_Paper(self, Index, Date, Title, Author, keyword, Page, URL):
        index = 1
        if Title.endswith('）') and keyword == '':
            text = Title.split('（')
            Title = text[0]
            # 部分标题内容重复，后带主题
            Flag1 = True
        for index in range(2, 99):
            InsertSQL = f"INSERT INTO [{self.contentTableName}] ([PaperIndex],[Date],[PaperTitle],[Author],[Keyword],[Page],[PaperURL]) VALUES ('{Index}','{Date}','{Title}','{Author}','{keyword}','{Page}','{URL}')"
            try:
                self.cursor.execute(InsertSQL)
                break
            except Exception as err:
                # 判断错误是否因主键冲突
                PRIMARYKEY_ERROR_CHECK = re.search('PRIMARY KEY', str(err))
                if PRIMARYKEY_ERROR_CHECK == None:
                    raise
                else:
                    URLSQL = f"SELECT [PaperURL] FROM [{self.contentTableName}] WHERE [Date] = '{Date}' AND [PaperTitle] = '{Title}'"
                    self.cursor.execute(URLSQL)
                    CollisionURL = self.cursor.fetchall()[0][0]
                    if Title == "图片报道" and URL != CollisionURL:
                        Title = f"图片报道{index}"
                        PrintAndSave(
                            f"标题修改提示：已将【{Index}_{Date}_{Title}_{Page}】的标题修改为【{Title}】")
                    else:
                        PageSQL = f"SELECT [Page],[PaperTitle],[Author],[Keyword] FROM [{self.contentTableName}] WHERE [Date] = '{Date}' AND [PaperTitle] = '{Title}'"
                        self.cursor.execute(PageSQL)
                        result = self.cursor.fetchall()
                        CollisionPage = result[0][0]
                        if Page in CollisionPage:
                            if Flag1:
                                self.UpdatePaperNumbers(Date)
                            PrintAndSave(
                                f"主键错误：\t【{Index}_{Date}_{Title}_{Page}】  \t已存在")
                        else:
                            NewPage = f"{CollisionPage}、{Page}"
                            UpdatePageSQL = f"UPDATE [{self.contentTableName}] SET [Page] = '{NewPage}' WHERE [Date] = '{Date}' AND [PaperTitle] = '{Title}'"
                            self.cursor.execute(UpdatePageSQL)
                            self.UpdatePaperNumbers(Date)
                            PrintAndSave(
                                f"版面修改提示：已将{Index}_{Date}_{Title}的版面从【{CollisionPage}】修改为【{NewPage}】")
                        break

    def UpdatePaperNumbers(self, Date):
        UpdatePaperNumbersSQL = f"UPDATE [{self.DateTableName}] SET [PaperNumbers] = [PaperNumbers] - 1 WHERE [Date] = '{Date}'"
        self.cursor.execute(UpdatePaperNumbersSQL)

    def DateCheck(self, date):
        date = date.replace('.', '-')
        DateSelectSQL = f"SELECT Date FROM [{self.DateTableName}] Where Date = '{date}'"
        DateList = self.cursor.execute(DateSelectSQL).fetchall()
        if bool(DateList):
            DateCompareSQL = f"""SELECT CASE WHEN [PaperNumbers] = [Exported-PaperNumbers] THEN 1 ELSE 0 END AS Compared
                                ,[PaperNumbers],[Exported-PaperNumbers] FROM dbo.[Date_Paper] WHERE [Date] = '{date}'""".replace("\n", " ")
            DateCompare = self.cursor.execute(DateCompareSQL).fetchall()
            if bool(DateCompare[0][0]):
                return False
        return True

    def ExportToXlsx():
        pass


class Xlsx:
    def __init__(self, filename):
        self.filename = filename
        self.WorkBook = openpyxl.Workbook()
        self.WorkSheet = self.WorkBook.active

    def CellStyle(self, xlsx):
        titlestyle = NamedStyle(name="TitleStyle",
                                font=Font(name='微软雅黑', size=24, bold=True, italic=False,
                                          vertAlign=None, underline='none', strike=False, color="00FFFFFF"),
                                fill=PatternFill(
                                    fill_type="solid", fgColor="000066CC"),
                                border=Border(),
                                alignment=Alignment(horizontal='center', vertical='center',
                                                    text_rotation=0, wrap_text=False, shrink_to_fit=True, indent=0),
                                number_format='General',
                                )
        oddstyle = NamedStyle(name="OddStyle",
                              font=Font(name='微软雅黑', size=12, bold=False, italic=False,
                                        vertAlign=None, underline='none', strike=False, color="00000000"),
                              fill=PatternFill(
                                  fill_type="solid", fgColor="00FFFFFF"),
                              border=Border(),
                              alignment=Alignment(horizontal='center', vertical='center',
                                                  text_rotation=0, wrap_text=False, shrink_to_fit=True, indent=0),
                              number_format='General',
                              )
        evenstyle = NamedStyle(name="EvenStyle",
                               font=Font(name='微软雅黑', size=12, bold=False, italic=False,
                                         vertAlign=None, underline='none', strike=False, color="00000000"),
                               fill=PatternFill(
                                   fill_type="solid", fgColor="00FFFFCC"),
                               border=Border(),
                               alignment=Alignment(horizontal='center', vertical='center',
                                                   text_rotation=0, wrap_text=False, shrink_to_fit=True, indent=0),
                               number_format='General',
                               )
        if not ('EvenStyle' in list(xlsx.style_names)):
            xlsx.add_named_style(evenstyle)
        if not ('TitleStyle' in list(xlsx.style_names)):
            xlsx.add_named_style(titlestyle)
        if not ('OddStyle' in list(xlsx.style_names)):
            xlsx.add_named_style(oddstyle)

    def SetCellStyle(self, Row):
        for Column in range(1, self.WorkSheet.max_column + 1):
            if Row % 2 == 0:
                self.WorkSheet[openpyxl.get_column_letter(
                    Column)+str(Row)].style = "OddStyle"
            else:
                self.WorkSheet[openpyxl.get_column_letter(
                    Column)+str(Row)].style = "EvenStyle"

    def SetColumnWidth(self, Column, Width):
        self.WorkSheet.column_dimensions[openpyxl.get_column_letter(
            Column)].width = Width


class DateList:

    TargetTitle = "NewspaperTitle(人民日报) _超星发现系统"
    Database = Database()

    def __init__(self, Mode="A") -> None:
        PrintAndSave(f"__________{Mode}_{time.time()}__________")
        if Mode == "A":
            self.driver = None
            self.PeopleDailyURL = "https://ss.zhizhen.com/s?sw=NewspaperTitle(%E4%BA%BA%E6%B0%91%E6%97%A5%E6%8A%A5)&nps=a5074152ece3d23811d0255ed743ac43"
            soup = self.Request()
            self.GetDateIndex(soup)
            self.GetDay("A")
        elif Mode == "B":
            self.PeopleDailyURL = "https://vpncas.ahut.edu.cn/https/77726476706e69737468656265737421e7e056d23d38614a760d87e29b5a2e/s?sw=NewspaperTitle%28%E4%BA%BA%E6%B0%91%E6%97%A5%E6%8A%A5%29&nps=a5074152ece3d23811d0255ed743ac43"
            with CreateEdgeDriverService() as driver:
                self.driver = driver
                soup = self.RequestSelenium(self.driver)
                self.GetDateIndex(soup)
                self.GetDay("B")

    def GetDateIndex(self, soup):
        DateListSource = soup.find_all("input", attrs={'id': 'guidedata1'})
        DateList = []
        if len(DateListSource) == 0:
            raise Exception("日期列表获取失败")
        for subDateList in DateListSource:
            DateString = subDateList.attrs["value"]
            pattern = re.compile(r'[1-2]\d{3}\.[0-1]\d\.[0-3]\d')
            DateList.append(pattern.findall(DateString))
        self.DateList = DateList

    def GetDay(self, Mode):
        for dates_Year in tqdm(self.DateList):
            for date in tqdm(dates_Year):
                if self.Database.DateCheck(date):
                    WebPage(self.Database, date,
                            self.PeopleDailyURL, Mode, self.driver)

    def Request(self):
        session = requests.Session()
        response = session.get(self.PeopleDailyURL)
        return BeautifulSoup(response.text, "html.parser")

    def RequestVPN(self):
        cookies = SmartLogin(self.PeopleDailyURL, smartUserName,
                             smartPassword, self.TargetTitle)
        # 获取首页
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'Connection': 'keep-alive',
            # 'Cookie': cookies,
            'Host': 'vpncas.ahut.edu.cn',
            'Referer': 'https://vpncas.ahut.edu.cn/https/77726476706e69737468656265737421f3f652d226387d44300d8db9d6562d/cas/login?service=https://vpncas.ahut.edu.cn/login?cas_login=true',
            'sec-ch-ua': '" Not;A Brand";v="99", "Microsoft Edge";v="103", "Chromium";v="103"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-Fetch-Dest': 'document',
            'sec-Fetch-Mode': 'navigate',
            'sec-Fetch-Site': 'same-origin',
            'sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.49'
        }
        session = requests.Session()
        # session.headers = headers
        response = session.get(self.PeopleDailyURL,
                               headers=headers, cookies=cookies)
        return response

    def RequestSelenium(self):
        SmartLogin(self.PeopleDailyURL, smartUserName,
                   smartPassword, self.TargetTitle)
        content = self.driver.page_source.encode('utf-8')
        return BeautifulSoup(content, 'lxml')


class WebPage:
    PaperNumberPattern = re.compile(r'返回<span>(.+?)</span>结果')
    TimeDelay = 25

    def __init__(self, Database, Date, PeopleDailyBaseURL, Mode="A", driver=None):
        self.PageSize = 100
        self.driver = driver
        self.mode = Mode
        self.PeopleDailyBaseDateURL = PeopleDailyBaseURL + "&npdate="
        self.Database = Database
        self.date = Date
        self.LoadFristPage()
        self.MaxPage = int(self.PaperNumber / self.PageSize + 0.9999999)
        self.GetNextPage()

    def LoadFristPage(self):
        # 获取日期链接
        self.FirstDateURL = self.PeopleDailyBaseDateURL + self.date +  \
            f"&size={self.PageSize}&isort=0&x=0_476&pages=1"
        # 获取日期页面
        dateSoup = self.requestPage(self.FirstDateURL)
        PaperNumberContent = dateSoup.find_all(
            "div", attrs={'class': 'left'})[0]
        PaperNumber = GetRegular(
            self.PaperNumberPattern, PaperNumberContent).replace(",", "")
        self.PaperNumber = int(PaperNumber)
        self.Database.InsertData_Date(
            self.date, self.PaperNumber, self.FirstDateURL)
        self.GetPaperList(dateSoup)

    def GetPaperList(self, dateSoup):
        PaperList = dateSoup.find_all(
            "div", attrs={'class': 'savelist clearfix'})
        for index, subPaperList in enumerate(PaperList):
            Paper(self.Database, index, self.date,
                  subPaperList, len(str(self.PaperNumber)))

    def GetNextPage(self):
        if self.MaxPage > 1:
            for PageIndex in range(2, self.MaxPage + 1):
                DateURL = self.PeopleDailyBaseDateURL + self.date + \
                    f"&size={self.PageSize}&isort=0&x=0_476&pages={PageIndex}"
                dateSoup = self.requestPage(DateURL)
                self.GetPaperList(dateSoup)

    def requestPage(self, URL):
        if self.mode == "A":
            while True:
                try:
                    response = requests.get(URL)
                    break
                except:
                    pass
            time.sleep(self.TimeDelay)
            soup = BeautifulSoup(response.text, "html.parser")
        elif self.mode == "B":
            self.driver.get(URL)
            self.driver.implicitly_wait(10)
            content = self.driver.page_source.encode('utf-8')
            soup = BeautifulSoup(content, "lxml")
        if "提示页面" in soup.text:
            raise Exception("需要验证码")
        return soup


class Paper:
    TitlePattern = re.compile(
        r'<input id="favtitle\d+" type="hidden" value="(.+?)"[>|/>]')
    URLPattern = re.compile(
        r'<input id="favurl\d+" type="hidden" value="(http://ss\.zhizhen\.com/.+?)"[>|/>]')
    AuthorPattern = re.compile(
        r'<input id="favauthor\d+" type="hidden" value="(.+?)"[>|/>]')
    AuthorRemainPattern = re.compile(
        r'"/>, <input id="(.+)?" type="hidden" value="')
    KeywordPattern = re.compile(
        r'<li>关键词：(.+?)</li>')
    PagePattern = re.compile(
        r'<li>出处：[\d\D]+人民日报[\w\W]+?(\d\d版.*|\d\d版：.*|\d\d版:.*).*</li>')

    def __init__(self, Database, index, date, subPaperList, IndexWeith):
        self.Database = Database
        self.index = index
        self.date = date
        self.subPaperList = subPaperList
        self.IndexWeith = IndexWeith
        self.InsertToDatebase()

    def InsertToDatebase(self):
        PaperForm = self.subPaperList.find_all(
            "form")[0].find_all(
            "input")
        PaperTitle = GetRegular(
            self.TitlePattern, PaperForm)
        PaperAuthor = GetRegular(
            self.AuthorPattern, PaperForm)
        if GetRegular(self.AuthorRemainPattern, PaperAuthor) != "":
            PaperAuthor = ""
        PaperURL = GetRegular(self.URLPattern, PaperForm)
        Paperul = self.subPaperList.find_all("ul")[0]
        PaperKeyword = GetRegular(
            self.KeywordPattern, Paperul).replace('<font color="Red">', "").replace('</font>', "")
        PaperPage = GetRegular(
            self.PagePattern, Paperul).replace("\xa0", "").replace(":", "：")
        if PaperPage != "":
            if PaperPage[0] != "第":
                PaperPage = "第" + PaperPage
        PaperIndex = self.date.replace(
            ".", "") + "_" + str((self.index + 1)).zfill(self.IndexWeith)
        self.Database.InsertData_Paper(
            PaperIndex, self.date, PaperTitle, PaperAuthor, PaperKeyword, PaperPage, PaperURL)


def GetRegular(pattern, text):
    result = pattern.search(str(text))
    if result:
        return result.group(1)
    else:
        return ""


def PrintAndSave(TEXT):
    print(TEXT)
    with open("log.txt", "a", encoding='utf8') as f:
        f.write(TEXT + "\n")


def main():
    if Mode == "A":
        DateList()
    if Mode == "B":
        while True:
            try:
                DateList("B")
            except:
                pass


if __name__ == '__main__':
    main()
